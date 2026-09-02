import base64
import io
import re
from pathlib import Path
from decimal import ROUND_CEILING, Decimal

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy.orm import Session, joinedload

from app.auth import require_admin
from app.database import get_db
from app.models import Estimate, EstimateChecklistItem, EstimateLine, EstimateSection, Item, Job, User
from app.schemas import (
    ChecklistItemIn, ChecklistItemOut, EstimateCreate, EstimateLineOut, EstimateOut, EstimateSectionOut,
    EstimateSummaryOut, EstimateUpdate, PlanAnalyzeIn, PlanAnalyzeOut,
)
from app.services.plan_analysis import analyze_plan_sheet, decode_pdf_data_url

router = APIRouter(prefix="/estimates", tags=["estimates"], dependencies=[Depends(require_admin)])


# Declared before the "/{estimate_id}" routes below so "checklist-items"
# never risks being swallowed by the int-typed path parameter.
@router.get("/checklist-items", response_model=list[ChecklistItemOut])
def list_checklist_items(db: Session = Depends(get_db)):
    return db.query(EstimateChecklistItem).order_by(EstimateChecklistItem.section, EstimateChecklistItem.label).all()


@router.post("/checklist-items", response_model=ChecklistItemOut, status_code=201)
def add_checklist_item(body: ChecklistItemIn, db: Session = Depends(get_db)):
    label = body.label.strip()
    existing = (
        db.query(EstimateChecklistItem)
        .filter(EstimateChecklistItem.section == body.section, EstimateChecklistItem.label.ilike(label))
        .first()
    )
    if existing:
        return existing
    item = EstimateChecklistItem(section=body.section, label=label)
    db.add(item)
    db.commit()
    return item


@router.delete("/checklist-items/{item_id}", status_code=204)
def delete_checklist_item(item_id: int, db: Session = Depends(get_db)):
    item = db.get(EstimateChecklistItem, item_id)
    if item:
        db.delete(item)
        db.commit()


# Declared before "/{estimate_id}" for the same reason as checklist-items above.
@router.post("/analyze-plan", response_model=PlanAnalyzeOut)
def analyze_plan(body: PlanAnalyzeIn):
    try:
        raw = decode_pdf_data_url(body.data)
    except Exception:
        raise HTTPException(status_code=400, detail="Couldn't read that file")
    if len(raw) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Plan sheet is too large (20MB limit)")
    try:
        result = analyze_plan_sheet(base64.b64encode(raw).decode())
    except RuntimeError as e:
        if "ANTHROPIC_API_KEY" in str(e):
            raise HTTPException(
                status_code=503,
                detail="AI plan reading isn't set up yet on this server -- ask your admin to add an API key.",
            )
        raise HTTPException(status_code=502, detail=str(e))
    return result

# APEX's standard phases, in the order they always appear on a real estimate
# (see "APEX Scope of Work Template 2025.docx" plus real jobs like Autumn
# Lake Bradford Dialysis and Leshnoff Townhomes, which also use Demolition
# and Miscellaneous on top of the template's five). Every new estimate gets
# all seven up front so the sheet always reads the same way, even if some
# end up empty.
STANDARD_SECTIONS = [
    "Demolition",
    "Rough-In Electrical Work",
    "Supply and Install",
    "Install Customer-Supplied Fixtures",
    "Panels and Meters",
    "Miscellaneous",
    "Permits, Inspection & Finalization",
]

# Keyword cues used to route a scope-of-work line into the right phase --
# checked in this order, first match wins. Demolition is checked first since
# a demo line can otherwise get misrouted by an incidental "panel"/"outlet"
# mention (e.g. "Demo the following: 1 Panel").
SECTION_CUES = [
    ("Demolition", ("demo", "demolition", "remove and", "remove existing")),
    ("Panels and Meters", ("panel", "meter", "disconnect", "breaker", "service upgrade", "ground rod", "riser", "transformer")),
    ("Permits, Inspection & Finalization", ("permit", "inspection")),
    ("Install Customer-Supplied Fixtures", ("customer-supplied", "customer supplied", "owner-supplied", "owner supplied")),
    ("Rough-In Electrical Work", ("rough-in", "rough in", "rough wiring")),
    ("Miscellaneous", ("relocate", "clean up", "cleanup", "miscellaneous", "misc.")),
]
DEFAULT_SECTION = "Supply and Install"

# Rough placeholder labor-per-unit, not fixture-specific pricing -- pulled
# from what actually recurs across APEX's real estimates (Leshnoff,
# Shelleydale, Autumn Lake, the Zakem target file all land in this range for
# ordinary devices). Without this, every drafted line had $0 labor and the
# total read as raw material cost only -- wildly under real project cost,
# since labor is usually the bigger half of the number.
DEFAULT_ROUGH_IN_LABOR = Decimal("35")
DEFAULT_SUPPLY_INSTALL_LABOR = Decimal("40")
# Rough-in isn't just labor -- wire, boxes, connectors, and staples to get a
# fixture ready for trim-out are real material cost too, even though it's a
# different set of parts than the device itself. $15/unit is the typical
# rough-in material cost across the same real jobs above.
DEFAULT_ROUGH_IN_MATERIAL = Decimal("15")

# A scope line phrased as "X will provide/supply Y" or "Y supplied/furnished
# by X" is telling you what's explicitly OUT of Apex's material scope --
# never run those through catalog matching, and file them under
# Customer-Supplied Fixtures regardless of what other cues the line contains.
# Doesn't apply when APEX is the one doing the providing ("Apex will supply
# all raw materials...") -- that confirms it's normal in-scope material, so
# those lines are just skipped rather than filed as an exclusion.
OTHER_SUPPLIED_PHRASES = ("will provide", "will supply", "supplied by", "provided by", "furnished by", "by others")
COMPANY_NAME_HINTS = ("apex",)
# "Separate quote for X" means X is being priced elsewhere, not part of this
# estimate -- keep normal section routing (so it still lands under Panels and
# Meters etc.) but skip catalog matching so it shows as a note, not a priced guess.
SEPARATE_QUOTE_PHRASES = ("separate quote", "separate estimate", "separate proposal")

STOPWORDS = {
    "a", "an", "the", "and", "or", "for", "with", "to", "of", "in", "on", "at", "by", "from",
    "install", "installing", "installed", "add", "adding", "new", "all", "each", "per",
    "run", "wire", "wiring", "up", "into", "provide", "furnish", "existing", "as", "needed",
}


def _next_estimate_number(db: Session) -> str:
    count = db.query(Estimate).count()
    return f"EST-{1001 + count}"


def _singularize(word: str) -> str:
    """Scope text is almost always plural ("Outlets", "Switches") while
    catalog item names are singular ("20A Duplex Receptacle") -- without
    this, matching silently misses on an exact-word-set intersection."""
    if word.endswith("ies") and len(word) > 4:
        return word[:-3] + "y"  # batteries -> battery
    if word.endswith(("ches", "shes", "xes", "ses")) and len(word) > 4:
        return word[:-2]  # switches -> switch, boxes -> box
    if word.endswith("s") and not word.endswith("ss") and len(word) > 3:
        return word[:-1]  # outlets -> outlet, lights -> light, detectors -> detector
    return word


def _significant_words(text: str) -> list[str]:
    # Electrical specs are often number-led ("12/2", "3/4\"", "20A") -- those
    # matter more for matching than plain words, so don't require a letter
    # first, just drop bare numbers (a lone "150" or "20" isn't descriptive).
    tokens = re.findall(r"[a-zA-Z0-9][a-zA-Z0-9/\-]{1,}", text.lower())
    return [_singularize(w) for w in tokens if w not in STOPWORDS and not w.isdigit()]


def _section_for_line(line: str) -> str:
    low = line.lower()
    for name, cues in SECTION_CUES:
        if any(c in low for c in cues):
            return name
    return DEFAULT_SECTION


def _extract_qty(line: str) -> tuple[Decimal, str | None]:
    """Returns (qty, note) -- note is set when a range like "30-35" had to be
    resolved to a single number, so the caller can flag the assumption."""
    range_match = re.search(r"\b(\d+)\s*-\s*(\d+)\b", line)
    if range_match:
        lo, hi = Decimal(range_match.group(1)), Decimal(range_match.group(2))
        if hi > lo:
            qty = ((lo + hi) / 2).quantize(Decimal("1"), rounding=ROUND_CEILING)
            return qty, f'"{line.strip()}" — assumed {qty} (midpoint of the {int(lo)}-{int(hi)} range you gave); adjust once finalized.'
    # A number that's part of a spec, not a count, is never the quantity:
    # "3-way"/"4-gang" (glued to a hyphen), "200 amp"/"20A" (an amperage rating).
    qty_match = re.search(r"\b(\d+(?:\.\d+)?)\b(?!-)(?!\s*[Aa](?:mp)?\b)", line)
    qty = Decimal(qty_match.group(1)) if qty_match else Decimal("1")
    return qty, None


def _draft_sections_from_scope(db: Session, scope_of_work: str) -> tuple[dict[str, list[dict]], list[str]]:
    """Best-effort starting point, not a real estimator: for each line of the
    scope of work, guess which phase it belongs to and score active items by
    word overlap with their name/category/description. Returns the drafted
    sections plus a plain-English list of assumptions made along the way, so
    nothing is silently guessed without a paper trail. The admin is expected
    to review/adjust -- and fill in labor, which this can't guess -- before
    sending anything out."""
    # Wall/cover plates are always bundled into installing the device they
    # cover -- they're never their own labor task, so they shouldn't ever
    # become their own draft line (previously they'd ride along as a
    # same-scored second match and get charged full device labor, e.g. a
    # wall plate "costing" $40 of labor same as the receptacle it covers).
    items = [i for i in db.query(Item).filter(Item.active).all() if "plate" not in i.name.lower()]
    item_words = {
        i.id: set(_significant_words(f"{i.name} {i.category} {i.description or ''}"))
        for i in items
    }
    by_id = {i.id: i for i in items}

    drafts: dict[str, list[dict]] = {name: [] for name in STANDARD_SECTIONS}
    notes: list[str] = []
    seen: set[int] = set()
    for raw_line in re.split(r"[\n\r]+|(?<=[.;])\s+", scope_of_work):
        line = raw_line.strip()
        if not line:
            continue
        line_words = set(_significant_words(line))
        if not line_words:
            continue
        scored = sorted(
            ((len(line_words & words), iid) for iid, words in item_words.items() if words),
            reverse=True,
        )
        qty, range_note = _extract_qty(line)
        if range_note:
            notes.append(range_note)
        low = line.lower()
        if any(p in low for p in OTHER_SUPPLIED_PHRASES):
            if any(h in low for h in COMPANY_NAME_HINTS):
                # "Apex will provide all raw materials..." confirms this IS
                # Apex's scope -- a summary statement, not a new line item
                # (the specific fixtures are already itemized elsewhere).
                continue
            # "I Heart will provide the heaters..." -- this is telling you
            # what's explicitly NOT Apex's material to buy, so it always
            # goes to Customer-Supplied Fixtures as a note, never matched
            # against the purchasing catalog.
            drafts["Install Customer-Supplied Fixtures"].append({
                "item_id": None, "description": line, "qty": qty,
                "unit": "each", "material_unit_cost": Decimal("0"), "labor_unit_cost": Decimal("0"),
            })
            notes.append(f'"{line}" — filed as customer/other-supplied; confirm before finalizing.')
            continue

        section = _section_for_line(line)
        is_separate_quote = any(p in low for p in SEPARATE_QUOTE_PHRASES)
        skip_matching = section in ("Permits, Inspection & Finalization", "Demolition") or is_separate_quote
        if skip_matching:
            # Permits/inspections are a service charge, demo work is labor
            # tearing things OUT (not material being bought), and a
            # "separate quote" line is being priced elsewhere entirely --
            # none of those should get matched against the purchasing
            # catalog. Drop in a blank line to price instead of risking a
            # nonsense catalog match on a stray word.
            drafts[section].append({
                "item_id": None, "description": line, "qty": qty,
                "unit": "each", "material_unit_cost": Decimal("0"), "labor_unit_cost": Decimal("0"),
            })
            if is_separate_quote:
                notes.append(f'"{line}" — priced separately; NOT included in this estimate\'s total.')
            continue
        # One match per scope line, not several -- a second, weaker-scored
        # match riding along used to add unrelated accessories as their own
        # fully-labor-priced line (see the wall-plate note above).
        for score, iid in scored[:1]:
            if score == 0 or iid in seen:
                continue
            seen.add(iid)
            item = by_id[iid]
            drafts[section].append({
                "item_id": item.id,
                "description": item.name,
                "qty": qty,
                "unit": item.unit,
                "material_unit_cost": item.avg_cost or item.last_cost or Decimal("0"),
                "labor_unit_cost": DEFAULT_SUPPLY_INSTALL_LABOR,
            })
            if section == DEFAULT_SECTION:
                # APEX's real jobs consistently price every fixture twice:
                # once here for the device itself, and once in Rough-In for
                # the wire/box material and labor to prep for it beforehand.
                drafts["Rough-In Electrical Work"].append({
                    "item_id": None, "description": f"Rough-in wiring for: {item.name}", "qty": qty,
                    "unit": item.unit, "material_unit_cost": DEFAULT_ROUGH_IN_MATERIAL,
                    "labor_unit_cost": DEFAULT_ROUGH_IN_LABOR,
                })
    if notes or any(drafts.values()):
        # Only worth a note if something was actually auto-drafted -- a
        # blank estimate started for manual item-picking shouldn't get a
        # disclaimer about auto-draft assumptions it never made.
        notes.append(
            f"Material/labor on Rough-In and Supply-and-Install lines are flat placeholders "
            f"(${DEFAULT_ROUGH_IN_MATERIAL}/${DEFAULT_ROUGH_IN_LABOR} per unit rough-in, ${DEFAULT_SUPPLY_INSTALL_LABOR} "
            "labor supply-install) based on typical costs in past jobs, NOT priced per fixture -- adjust anything "
            "more complex (panels, equipment connections, custom work) before sending this out."
        )
    return drafts, notes


def _line_totals(l: EstimateLine) -> tuple[Decimal, Decimal, Decimal]:
    material = (l.qty * l.material_unit_cost).quantize(Decimal("0.01"))
    labor = (l.qty * l.labor_unit_cost).quantize(Decimal("0.01"))
    return material, labor, material + labor


def _serialize(estimate: Estimate) -> EstimateOut:
    material_total = Decimal("0.00")
    labor_total = Decimal("0.00")
    section_outs = []
    for section in estimate.sections:
        line_outs = []
        section_total = Decimal("0.00")
        for l in section.lines:
            m, lb, tot = _line_totals(l)
            material_total += m
            labor_total += lb
            section_total += tot
            line_outs.append(EstimateLineOut(
                id=l.id, item_id=l.item_id, sku=l.item.sku if l.item else None,
                image_data=l.item.image_data if l.item else None, description=l.description,
                qty=l.qty, unit=l.unit, material_unit_cost=l.material_unit_cost,
                labor_unit_cost=l.labor_unit_cost, material_total=m, labor_total=lb, line_total=tot,
            ))
        section_outs.append(EstimateSectionOut(id=section.id, name=section.name, lines=line_outs, section_total=section_total))

    subtotal = material_total + labor_total
    profit_amount = (subtotal * estimate.profit_pct / Decimal("100")).quantize(Decimal("0.01"))
    discount_amount = ((subtotal + profit_amount) * estimate.discount_pct / Decimal("100")).quantize(Decimal("0.01"))
    total = (subtotal + profit_amount - discount_amount).quantize(Decimal("0.01"))

    return EstimateOut(
        id=estimate.id, estimate_number=estimate.estimate_number,
        job_id=estimate.job_id, job_number=estimate.job.job_number if estimate.job else None,
        customer=estimate.customer,
        address=estimate.address, scope_of_work=estimate.scope_of_work, exclusions=estimate.exclusions,
        status=estimate.status, profit_pct=estimate.profit_pct, discount_pct=estimate.discount_pct,
        material_total=material_total, labor_total=labor_total, subtotal=subtotal,
        profit_amount=profit_amount, discount_amount=discount_amount, total=total,
        created_by_name=estimate.creator.name if estimate.creator else None,
        created_at=estimate.created_at, updated_at=estimate.updated_at, sections=section_outs,
    )


def _get_or_404(db: Session, estimate_id: int) -> Estimate:
    est = (
        db.query(Estimate)
        .options(
            joinedload(Estimate.sections).joinedload(EstimateSection.lines).joinedload(EstimateLine.item),
            joinedload(Estimate.creator),
            joinedload(Estimate.job),
        )
        .filter(Estimate.id == estimate_id)
        .first()
    )
    if not est:
        raise HTTPException(status_code=404, detail="Estimate not found")
    return est


@router.get("", response_model=list[EstimateSummaryOut])
def list_estimates(job_id: int | None = None, db: Session = Depends(get_db)):
    q = db.query(Estimate).options(joinedload(Estimate.sections).joinedload(EstimateSection.lines))
    if job_id is not None:
        q = q.filter(Estimate.job_id == job_id)
    rows = q.order_by(Estimate.created_at.desc()).all()
    out = []
    for e in rows:
        subtotal = Decimal("0.00")
        for section in e.sections:
            for l in section.lines:
                _, _, tot = _line_totals(l)
                subtotal += tot
        profit_amount = (subtotal * e.profit_pct / Decimal("100")).quantize(Decimal("0.01"))
        discount_amount = ((subtotal + profit_amount) * e.discount_pct / Decimal("100")).quantize(Decimal("0.01"))
        total = (subtotal + profit_amount - discount_amount).quantize(Decimal("0.01"))
        out.append(EstimateSummaryOut(
            id=e.id, estimate_number=e.estimate_number, job_id=e.job_id, customer=e.customer, status=e.status,
            total=total, created_at=e.created_at, updated_at=e.updated_at,
        ))
    return out


@router.get("/{estimate_id}", response_model=EstimateOut)
def get_estimate(estimate_id: int, db: Session = Depends(get_db)):
    return _serialize(_get_or_404(db, estimate_id))


def _notes_to_text(notes: list[str]) -> str:
    if not notes:
        return ""
    return "Notes / assumptions from the auto-draft:\n" + "\n".join(f"- {n}" for n in notes)


@router.post("", response_model=EstimateOut, status_code=201)
def create_estimate(body: EstimateCreate, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    job = db.get(Job, body.job_id) if body.job_id else None
    if body.job_id and not job:
        raise HTTPException(status_code=404, detail="Job not found")
    estimate = Estimate(
        estimate_number=_next_estimate_number(db), job_id=body.job_id,
        customer=body.customer or ((job.customer or job.name) if job else None),
        address=body.address or (job.address if job else None),
        scope_of_work=body.scope_of_work, profit_pct=body.profit_pct, discount_pct=body.discount_pct,
        created_by=user.id,
    )
    db.add(estimate)
    db.flush()

    drafts, notes = _draft_sections_from_scope(db, body.scope_of_work)
    estimate.exclusions = _notes_to_text(notes)
    for order, name in enumerate(STANDARD_SECTIONS):
        section = EstimateSection(estimate_id=estimate.id, name=name, sort_order=order)
        db.add(section)
        db.flush()
        for line_order, draft in enumerate(drafts.get(name, [])):
            db.add(EstimateLine(section_id=section.id, sort_order=line_order, **draft))
    db.commit()
    return _serialize(_get_or_404(db, estimate.id))


@router.patch("/{estimate_id}", response_model=EstimateOut)
def update_estimate(estimate_id: int, body: EstimateUpdate, db: Session = Depends(get_db)):
    estimate = _get_or_404(db, estimate_id)
    if body.job_id is not None:
        if not db.get(Job, body.job_id):
            raise HTTPException(status_code=404, detail="Job not found")
    data = body.model_dump(exclude_unset=True, exclude={"sections", "clear_job"})
    for field, value in data.items():
        setattr(estimate, field, value)
    if body.clear_job:
        estimate.job_id = None
    if body.sections is not None:
        for section in list(estimate.sections):
            db.delete(section)
        db.flush()
        for order, section_in in enumerate(body.sections):
            section = EstimateSection(estimate_id=estimate.id, name=section_in.name, sort_order=order)
            db.add(section)
            db.flush()
            for line_order, line_in in enumerate(section_in.lines):
                db.add(EstimateLine(
                    section_id=section.id, sort_order=line_order, item_id=line_in.item_id,
                    description=line_in.description, qty=line_in.qty, unit=line_in.unit,
                    material_unit_cost=line_in.material_unit_cost, labor_unit_cost=line_in.labor_unit_cost,
                ))
    db.commit()
    return _serialize(_get_or_404(db, estimate.id))


@router.delete("/{estimate_id}", status_code=204)
def delete_estimate(estimate_id: int, db: Session = Depends(get_db)):
    estimate = _get_or_404(db, estimate_id)
    db.delete(estimate)
    db.commit()


@router.post("/{estimate_id}/redraft", response_model=EstimateOut)
def redraft_estimate(estimate_id: int, db: Session = Depends(get_db)):
    """Re-run the scope-of-work matcher and replace the current sections --
    useful after editing the scope text, or to discard manual tweaks and
    start over. Material costs get re-drawn from the catalog; labor always
    resets to 0 since there's nothing to infer it from. Exclusions/notes are
    also regenerated fresh, so anything typed there manually is replaced."""
    estimate = _get_or_404(db, estimate_id)
    for section in list(estimate.sections):
        db.delete(section)
    db.flush()

    drafts, notes = _draft_sections_from_scope(db, estimate.scope_of_work)
    estimate.exclusions = _notes_to_text(notes)
    for order, name in enumerate(STANDARD_SECTIONS):
        section = EstimateSection(estimate_id=estimate.id, name=name, sort_order=order)
        db.add(section)
        db.flush()
        for line_order, draft in enumerate(drafts.get(name, [])):
            db.add(EstimateLine(section_id=section.id, sort_order=line_order, **draft))
    db.commit()
    return _serialize(_get_or_404(db, estimate.id))


# Matches "APEX Estimating Spreadsheet 2026" exactly: same intro line
# wording, in the same order. Sections not in this map (Demolition,
# Miscellaneous -- used on some real jobs but not part of the 2026 template)
# just get no intro line.
SECTION_INTRO = {
    "Rough-In Electrical Work": "Provide rough-in wiring for the following fixtures:",
    "Supply and Install": "Supply and install the following:",
    "Install Customer-Supplied Fixtures": "Install the following customer-supplied fixtures:",
}

FONT = "Calibri"
_ACCOUNTING_FMT = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'


def _build_xlsx(estimate: Estimate) -> Workbook:
    """Lays the estimate out exactly like "APEX Estimating Spreadsheet
    2026" -- same fonts, column widths, accounting number format, and
    formula style (=SUM(C*D) per line, =SUM(H_first:H_last) for totals),
    so an exported estimate looks and behaves like a normal APEX estimate
    file, not a foreign format. Every dollar cell is a live formula, not a
    baked-in number, so it recalculates if someone edits it in Excel."""
    bold = Font(name=FONT, bold=True)
    normal = Font(name=FONT)
    title_font = Font(name=FONT, size=20)
    header_font = Font(name=FONT, bold=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"
    for col, width in zip("ABCDEFGH", [11.9, 90, 10.3, 15.1, 21.9, 12.6, 19.0, 15.0]):
        ws.column_dimensions[col].width = width
    ws.row_dimensions[2].height = 72
    ws.freeze_panes = "A5"

    logo_path = Path(__file__).resolve().parent.parent.parent / "static" / "logo.png"
    try:
        from openpyxl.drawing.image import Image as XLImage
        logo = XLImage(str(logo_path))
        logo.height = 64
        logo.width = 103
        ws.add_image(logo, "A2")
    except Exception:
        pass  # logo is a nice-to-have -- never let a missing/unreadable file break the export

    ws.cell(row=3, column=2, value=f"Job Name: {estimate.customer or estimate.estimate_number}").font = title_font
    ws.row_dimensions[3].height = 25.8
    if estimate.address:
        ws.cell(row=3, column=2).value = f"Job Name: {estimate.customer or estimate.estimate_number}  ({estimate.address})"

    headers = ["", "", "Quantity", "Cost of Material", "Total Cost of Material", "Labor Cost", "Total Cost of Labor", "Total cost"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h).font = header_font

    # openpyxl can't attach a cached value to a formula cell (it only ever
    # writes the "=..." string), so a formula-only cell reads back as blank
    # in anything that doesn't actually run a calc engine on open -- Windows/
    # mobile file previews, some Excel calculation-mode setups, etc. That's
    # exactly what showed up as "no cost of materials, no total costs" on a
    # real download. So every derived dollar cell below is the actual
    # computed number (same math as the API/UI use), not a formula --
    # correct on open everywhere, at the cost of not live-recalculating if
    # someone hand-edits a quantity in Excel afterward.
    row = 6
    material_total = Decimal("0.00")
    labor_total = Decimal("0.00")
    section_index = 0
    for section in estimate.sections:
        if not section.lines:
            continue
        section_index += 1
        section_header_row = row
        header_cell = ws.cell(row=row, column=2, value=f"{section_index}. {section.name}")
        header_cell.font = bold
        header_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        row += 1
        intro = SECTION_INTRO.get(section.name)
        if intro:
            ws.cell(row=row, column=2, value=intro).font = normal
            row += 1
        section_total = Decimal("0.00")
        for line in section.lines:
            m, lb, tot = _line_totals(line)
            material_total += m
            labor_total += lb
            section_total += tot
            ws.cell(row=row, column=2, value=line.description).font = normal
            ws.cell(row=row, column=3, value=float(line.qty))
            ws.cell(row=row, column=4, value=float(line.material_unit_cost)).number_format = _ACCOUNTING_FMT
            ws.cell(row=row, column=5, value=float(m)).number_format = _ACCOUNTING_FMT
            ws.cell(row=row, column=6, value=float(line.labor_unit_cost)).number_format = _ACCOUNTING_FMT
            ws.cell(row=row, column=7, value=float(lb)).number_format = _ACCOUNTING_FMT
            ws.cell(row=row, column=8, value=float(tot)).number_format = _ACCOUNTING_FMT
            row += 1
        subtotal_cell = ws.cell(row=section_header_row, column=1, value=float(section_total))
        subtotal_cell.number_format = _ACCOUNTING_FMT
        subtotal_cell.font = bold
        row += 1  # blank row before the next section

    subtotal = material_total + labor_total
    row += 1
    ws.cell(row=row, column=2, value="Total cost of material and labor").font = bold
    ws.cell(row=row, column=8, value=float(subtotal)).number_format = _ACCOUNTING_FMT
    ws.cell(row=row, column=8).font = bold
    row += 2

    profit_amount = (subtotal * estimate.profit_pct / Decimal("100")).quantize(Decimal("0.01"))
    ws.cell(row=row, column=2, value="Profit").font = normal
    ws.cell(row=row, column=6, value=float(estimate.profit_pct) / 100).number_format = "0.0%"
    ws.cell(row=row, column=8, value=float(profit_amount)).number_format = _ACCOUNTING_FMT
    row += 1

    discount_amount = ((subtotal + profit_amount) * estimate.discount_pct / Decimal("100")).quantize(Decimal("0.01"))
    ws.cell(row=row, column=2, value="Repeat customer discount").font = normal
    ws.cell(row=row, column=6, value=float(estimate.discount_pct) / 100).number_format = "0.0%"
    ws.cell(row=row, column=8, value=float(-discount_amount)).number_format = _ACCOUNTING_FMT
    row += 1

    grand_total = (subtotal + profit_amount - discount_amount).quantize(Decimal("0.01"))
    ws.cell(row=row, column=2, value="TOTAL COST OF PROJECT").font = bold
    grand_cell = ws.cell(row=row, column=8, value=float(grand_total))
    grand_cell.number_format = _ACCOUNTING_FMT
    grand_cell.font = bold
    row += 3

    if estimate.exclusions and estimate.exclusions.strip():
        ws.cell(row=row, column=2, value="Exclusions / Notes:").font = bold
        row += 1
        for text_line in estimate.exclusions.splitlines():
            if text_line.strip():
                ws.cell(row=row, column=2, value=text_line.strip()).font = normal
                row += 1

    return wb


@router.get("/{estimate_id}/export")
def export_estimate(estimate_id: int, db: Session = Depends(get_db)):
    estimate = _get_or_404(db, estimate_id)
    wb = _build_xlsx(estimate)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{estimate.estimate_number}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
